from __future__ import annotations

from io import BytesIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from backend.models import ProjectSnapshot, TaskCreate
from backend.scheduler import PlanValidationError, schedule

HEADERS = ["task", "description", "executor", "duration", "predecessors"]
RUSSIAN_HEADERS = ["задача", "описание", "исполнитель", "длительность", "предшественники"]
HEADER_ALIASES = dict(zip(RUSSIAN_HEADERS, HEADERS))
RUSSIAN_ASSIGNEES = {
    "Anna": "Анна",
    "Elena": "Елена",
    "Mikhail": "Михаил",
    "Daria": "Дарья",
    "Pavel": "Павел",
}
ASSIGNEE_ALIASES = {
    localized.casefold(): canonical for canonical, localized in RUSSIAN_ASSIGNEES.items()
}


def parse_excel(content: bytes, project_name: str, start_date) -> ProjectSnapshot:
    try:
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
    except Exception as exc:
        raise PlanValidationError("The file is not a readable .xlsx workbook") from exc
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise PlanValidationError("The workbook is empty")
    actual = [
        HEADER_ALIASES.get(str(value or "").strip().casefold(), str(value or "").strip().casefold())
        for value in rows[0]
    ]
    missing = [header for header in HEADERS if header not in actual]
    if missing:
        raise PlanValidationError(f"Missing columns: {', '.join(missing)}")
    positions = {name: actual.index(name) for name in HEADERS}
    raw = []
    for number, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        name = str(row[positions["task"]] or "").strip()
        if not name:
            raise PlanValidationError(f"Row {number}: task name is empty")
        try:
            duration = int(row[positions["duration"]])
        except (TypeError, ValueError) as exc:
            raise PlanValidationError(f"Row {number}: duration must be a whole number") from exc
        raw.append(
            {
                "id": str(uuid4()),
                "name": name,
                "description": str(row[positions["description"]] or "").strip(),
                "assignees": [
                    ASSIGNEE_ALIASES.get(name.strip().casefold(), name.strip())
                    for name in str(row[positions["executor"]] or "").replace(";", ",").split(",")
                    if name.strip()
                ],
                "duration": duration,
                "predecessors": [
                    part.strip()
                    for part in str(row[positions["predecessors"]] or "").split(",")
                    if part.strip()
                ],
            }
        )
    name_to_id = {item["name"].casefold(): item["id"] for item in raw}
    if len(name_to_id) != len(raw):
        raise PlanValidationError("Task names must be unique")
    tasks = []
    for item in raw:
        names = item.pop("predecessors")
        unknown = [name for name in names if name.casefold() not in name_to_id]
        if unknown:
            raise PlanValidationError(f"{item['name']}: unknown predecessor {unknown[0]}")
        tasks.append(
            TaskCreate(**item, predecessor_ids=[name_to_id[name.casefold()] for name in names])
        )
    snapshot = ProjectSnapshot(name=project_name, start_date=start_date, tasks=tasks)
    schedule(snapshot)
    return snapshot


def export_excel(project, language: str = "en") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    localized = language == "ru"
    sheet.title = "План" if localized else "Plan"
    sheet.append(RUSSIAN_HEADERS if localized else HEADERS)
    names = {task.id: task.name for task in project.tasks}
    for task in project.tasks:
        sheet.append(
            [
                task.name,
                task.description,
                ", ".join(
                    RUSSIAN_ASSIGNEES.get(assignee, assignee) if localized else assignee
                    for assignee in task.assignees
                ),
                task.duration,
                ", ".join(names[item] for item in task.predecessor_ids),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in zip("ABCDE", [30, 48, 20, 14, 36]):
        sheet.column_dimensions[column].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
