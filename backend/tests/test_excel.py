from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from backend.excel import export_excel, parse_excel
from backend.models import Person
from backend.scheduler import PlanValidationError
from backend.seed import seed_snapshot
from backend.store import ProjectStore

SAMPLE_DIR = Path(__file__).resolve().parents[2] / "sample" / "import-tests"


def workbook(rows, headers=None):
    book = Workbook()
    sheet = book.active
    sheet.append(headers or ["task", "description", "executor", "duration", "predecessors"])
    for row in rows:
        sheet.append(row)
    data = BytesIO()
    book.save(data)
    return data.getvalue()


def test_excel_round_trip():
    source = ProjectStore()
    source.snapshot.tasks[0].assignees = [Person.ELENA, Person.PAVEL]
    data = export_excel(source.project())
    parsed = parse_excel(data, "Imported", source.snapshot.start_date)
    assert len(parsed.tasks) == 12
    assert parsed.tasks[0].assignees == ["Elena", "Pavel"]
    assert parsed.tasks[-1].predecessor_ids


def test_russian_excel_round_trip():
    source = ProjectStore()
    source.snapshot.tasks[0].assignees = [Person.ELENA, Person.PAVEL]

    data = export_excel(source.project(), language="ru")
    sheet = load_workbook(BytesIO(data), read_only=True).active
    parsed = parse_excel(data, "Импортированный проект", source.snapshot.start_date)

    assert sheet.title == "План"
    assert [cell.value for cell in sheet[1]] == [
        "задача",
        "описание",
        "исполнитель",
        "длительность",
        "предшественники",
    ]
    assert sheet["C2"].value == "Елена, Павел"
    assert parsed.tasks[0].assignees == ["Elena", "Pavel"]


@pytest.mark.parametrize(
    ("filename", "expected_tasks"),
    [
        ("05-ru-small-linear-8-tasks.xlsx", 8),
        ("06-ru-medium-20-tasks.xlsx", 20),
        ("07-ru-large-complex-45-tasks.xlsx", 45),
    ],
)
def test_russian_sample_workbook_import(filename, expected_tasks):
    data = (SAMPLE_DIR / filename).read_bytes()

    parsed = parse_excel(data, "Импортированный проект", seed_snapshot().start_date)

    assert len(parsed.tasks) == expected_tasks
    assert all(task.assignees for task in parsed.tasks)


def test_excel_normalizes_common_header_variants():
    data = workbook(
        [["Проверка", "Описание", "Анна", 1, ""]],
        ["\ufefftask", "descriptions", "executors", "durations", "predecessor"],
    )

    parsed = parse_excel(data, "Импортированный проект", seed_snapshot().start_date)

    assert parsed.tasks[0].name == "Проверка"
    assert parsed.tasks[0].assignees == ["Anna"]


def test_excel_unknown_predecessor():
    with pytest.raises(PlanValidationError, match="unknown predecessor"):
        parse_excel(
            workbook([["Test", "", "Anna", 2, "Missing"]]), "Test", seed_snapshot().start_date
        )
